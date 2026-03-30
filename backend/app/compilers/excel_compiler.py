"""Excel Compiler - Compiles agent results into Excel files."""
import io
from typing import Dict, Any, List
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)


class ExcelCompiler:
    """Compiles agent outputs into Excel workbooks."""
    
    def compile(
        self,
        structure: Dict[str, Any],
        data: List[Dict[str, Any]],
        formulas: Dict[str, str],
        charts: List[Dict[str, Any]],
        styling: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Compile all components into an Excel file."""
        
        try:
            wb = Workbook()
            
            # Handle structure - can be list (multiple sheets) or dict (single sheet)
            sheets_config = structure if isinstance(structure, list) else [structure]
            
            # Create sheets
            for idx, sheet_config in enumerate(sheets_config):
                if idx == 0:
                    ws = wb.active
                    ws.title = sheet_config.get("sheet_name", "Data") if isinstance(sheet_config, dict) else sheet_config.name
                else:
                    ws = wb.create_sheet(title=sheet_config.get("sheet_name", f"Sheet{idx+1}") if isinstance(sheet_config, dict) else sheet_config.name)
                
                # Write data to sheet
                self._write_sheet_data(ws, sheet_config, data, formulas, styling)
            
            # Add dashboard sheet if charts exist
            if charts:
                dashboard = wb.create_sheet(title="Dashboard")
                self._create_dashboard(dashboard, charts, data)
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return {
                "content": output.getvalue(),
                "filename": "generated_spreadsheet.xlsx",
                "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "size": len(output.getvalue())
            }
            
        except Exception as e:
            logger.error(f"Excel compilation failed: {e}")
            # Return minimal valid workbook
            return self._create_fallback_output()
    
    def _write_sheet_data(
        self,
        ws,
        sheet_config: Dict[str, Any],
        data: List[Dict[str, Any]],
        formulas: Dict[str, str],
        styling: Dict[str, Any]
    ):
        """Write data to a worksheet."""
        
        # Get column definitions
        if isinstance(sheet_config, dict):
            columns = sheet_config.get("columns", [])
        else:
            columns = sheet_config.columns
        
        if not columns and data:
            # Infer columns from data
            columns = [{"name": k, "type": "text"} for k in data[0].keys()]
        
        if not columns:
            columns = [{"name": "Data", "type": "text"}]
        
        # Write headers
        header_style = styling.get("header_style", {})
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col.get("name", "Column"))
            self._apply_header_style(cell, header_style)
        
        # Write data
        data_style = styling.get("data_style", {})
        use_alternating = styling.get("alternating_rows", {}).get("enabled", False)
        alt_fill = styling.get("alternating_rows", {}).get("fill", {})
        
        data_rows = data[:100] if len(data) > 100 else data  # Limit to 100 rows for now
        
        for row_idx, row_data in enumerate(data_rows, 2):
            for col_idx, col in enumerate(columns, 1):
                col_name = col.get("name", "")
                value = row_data.get(col_name, "") if isinstance(row_data, dict) else row_data[col_idx - 1] if isinstance(row_data, (list, tuple)) else ""
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                self._apply_data_style(cell, data_style)
                
                # Apply alternating row color
                if use_alternating and row_idx % 2 == 0:
                    if alt_fill.get("fgColor", {}).get("rgb"):
                        cell.fill = PatternFill(
                            start_color=alt_fill["fgColor"]["rgb"],
                            end_color=alt_fill["fgColor"]["rgb"],
                            fill_type="solid"
                        )
        
        # Apply formulas
        for cell_ref, formula in formulas.items():
            try:
                # Parse cell reference (e.g., "A2" or "Sheet1!A2")
                if "!" in cell_ref:
                    sheet_ref, cell_addr = cell_ref.split("!")
                    if ws.title == sheet_ref:
                        ws[cell_addr] = formula
                else:
                    ws[cell_ref] = formula
            except Exception as e:
                logger.warning(f"Failed to apply formula {cell_ref}={formula}: {e}")
        
        # Auto-adjust column widths
        for col_idx, col in enumerate(columns, 1):
            col_letter = get_column_letter(col_idx)
            max_length = len(str(col.get("name", "")))
            
            for row in range(2, min(len(data_rows) + 2, 12)):  # Check first 10 rows
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
    
    def _create_dashboard(self, ws, charts: List[Dict[str, Any]], data: List[Dict[str, Any]]):
        """Create a dashboard sheet with charts."""
        
        ws.cell(row=1, column=1, value="Dashboard")
        ws.cell(row=1, column=1).font = Font(bold=True, size=16)
        
        # Create charts
        chart_idx = 0
        for chart_config in charts[:4]:  # Max 4 charts
            chart_type = chart_config.get("type", "column")
            
            if chart_type == "column":
                chart = BarChart()
                chart.type = chart_config.get("subtype", "col")
            elif chart_type == "bar":
                chart = BarChart()
                chart.type = chart_config.get("subtype", "bar")
            elif chart_type == "line":
                chart = LineChart()
            elif chart_type == "pie":
                chart = PieChart()
            else:
                chart = BarChart()
            
            chart.title = chart_config.get("title", "Chart")
            
            # Position chart
            row_offset = 3 + (chart_idx // 2) * 16
            col_offset = 1 if chart_idx % 2 == 0 else 8
            
            ws.add_chart(chart, f"{get_column_letter(col_offset)}{row_offset}")
            chart_idx += 1
    
    def _apply_header_style(self, cell, style: Dict[str, Any]):
        """Apply header styling to a cell."""
        font_config = style.get("font", {})
        cell.font = Font(
            bold=font_config.get("bold", True),
            size=font_config.get("size", 11),
            color=font_config.get("color", "FFFFFF")
        )
        
        fill_config = style.get("fill", {})
        if fill_config.get("fgColor", {}).get("rgb"):
            cell.fill = PatternFill(
                start_color=fill_config["fgColor"]["rgb"],
                end_color=fill_config["fgColor"]["rgb"],
                fill_type="solid"
            )
        
        alignment_config = style.get("alignment", {})
        cell.alignment = Alignment(
            horizontal=alignment_config.get("horizontal", "center"),
            vertical=alignment_config.get("vertical", "center")
        )
    
    def _apply_data_style(self, cell, style: Dict[str, Any]):
        """Apply data styling to a cell."""
        font_config = style.get("font", {})
        cell.font = Font(
            size=font_config.get("size", 10),
            name=font_config.get("name", "Calibri")
        )
        
        alignment_config = style.get("alignment", {})
        cell.alignment = Alignment(
            horizontal=alignment_config.get("horizontal", "left"),
            vertical=alignment_config.get("vertical", "center")
        )
    
    def _create_fallback_output(self) -> Dict[str, Any]:
        """Create a minimal fallback Excel file."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        ws["A1"] = "Error"
        ws["B1"] = "Document generation failed"
        ws["A2"] = "Please"
        ws["B2"] = "Try again with a different request"
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return {
            "content": output.getvalue(),
            "filename": "error.xlsx",
            "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "size": len(output.getvalue())
        }


# Global instance
excel_compiler = ExcelCompiler()
